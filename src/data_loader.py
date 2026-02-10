"""AR Email Automation - XLSX Data Loader.

Parses the NY Account Receivables Overdue workbook and returns typed
``Invoice`` and ``Contact`` objects ready for template rendering.

Supported data sources
~~~~~~~~~~~~~~~~~~~~~~
* **File path** -- local ``.xlsx`` file exported from Google Sheets.
* **Bytes buffer** -- ``io.BytesIO`` for future Google Sheets API or
  HTTP-download pipelines.

Sheet layout (see ``agent-outputs/04-xlsx-schema-analysis.md``):

+-----------------------+---------------------------------------------------+
| Sheet                 | Purpose                                           |
+=======================+===================================================+
| ``Overdue {date}``    | Weekly snapshot of overdue invoices (active)       |
| ``Data``              | Formula-driven live view (fallback)                |
| ``Managers``          | Retailer POC directory (620 records)               |
| ``balances``          | Full AR ledger with financial breakdown            |
| ``Territory``         | Sales-rep territory assignments                    |
+-----------------------+---------------------------------------------------+

Usage::

    from src.data_loader import load_workbook

    result = load_workbook("data/NY Account Receivables_Overdue.xlsx")
    print(f"Invoices: {len(result.invoices)}")
    print(f"Contacts: {len(result.contacts)}")
    result.print_summary()
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import IO, Optional, Union

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Contact, Invoice, InvoiceStatus, SkipReason, Tier

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Pattern for weekly snapshot sheet names: "Overdue 2-3", "Overdue 1-26", etc.
_OVERDUE_SHEET_PATTERN = re.compile(
    r"^Overdue\s+(\d{1,2})-(\d{1,2})$", re.IGNORECASE
)

_MANAGERS_SHEET = "Managers"

# Column header aliases -- mapped by *header text* so we are resilient
# to column reordering across weekly snapshots.
_OVERDUE_HEADERS: dict[str, list[str]] = {
    "order_no":       ["Order No", "Order Number"],
    "location":       ["Location"],
    "account_mgr":    ["Account Manager"],
    # Note: the source sheet has a typo -- "Acocunt"
    "am_phone":       [
        "Acocunt Manager Phone Number",
        "Account Manager Phone Number",
        "Account Manager Phone#",
    ],
    "email_sent":     ["Email Sent"],
    "caller":         ["Caller"],
    "called":         ["Called"],
    "made_contact":   ["Made Contact"],
    "due_date":       ["Due Date"],
    "days_over":      ["Days Over", "Days Overdue", "Overdue"],
    "total_due":      ["Total Due"],
    "rep":            ["Rep"],
    "paid":           ["Paid"],
    "follow_up_date": ["F/U Date"],
    "status":         ["Status"],
    "notes":          ["Notes"],
}

_MANAGERS_HEADERS: dict[str, list[str]] = {
    "retailer_name": ["Retailer Name (DBA)"],
    "account_mgr":   ["Account Manager"],
    "am_phone":      [
        "Account Manager Phone#",
        "Account Manager Phone Number",
        "Acocunt Manager Phone Number",
    ],
    "poc_name":      ["POC Name & Title"],
    "poc_email":     ["POC Email"],
    "poc_phone":     ["POC Phone"],
}

# Cell values that should be treated as null / unknown.
_NULL_SIGNALS: set[str | None] = {"", "#N/A", "N/A", "#REF!", None}

# Keywords indicating an AR-focused contact (for primary selection).
_AR_TITLE_KEYWORDS = ["ap", "accounting", "finance", "billing", "accounts payable"]


# ---------------------------------------------------------------------------
# Result container
# ---------------------------------------------------------------------------

@dataclass
class LoadResult:
    """Aggregated output from :func:`load_workbook`."""

    invoices: list[Invoice] = field(default_factory=list)
    contacts: list[Contact] = field(default_factory=list)
    contacts_by_name: dict[str, Contact] = field(default_factory=dict)

    # Metadata
    source_file: str | None = None
    overdue_sheet_used: str | None = None
    total_rows_scanned: int = 0
    empty_rows_skipped: int = 0
    warnings: list[str] = field(default_factory=list)

    # Matching diagnostics
    matched_locations: list[str] = field(default_factory=list)
    unmatched_locations: list[str] = field(default_factory=list)

    @property
    def match_rate(self) -> float:
        total = len(self.matched_locations) + len(self.unmatched_locations)
        return len(self.matched_locations) / total if total else 0.0

    @property
    def paid_invoices(self) -> list[Invoice]:
        return [i for i in self.invoices if i.paid]

    @property
    def unpaid_invoices(self) -> list[Invoice]:
        return [i for i in self.invoices if not i.paid]

    @property
    def actionable_invoices(self) -> list[Invoice]:
        return [i for i in self.invoices if i.is_sendable]

    @property
    def total_ar(self) -> float:
        return sum(i.amount for i in self.invoices)

    @property
    def unpaid_ar(self) -> float:
        return sum(i.amount for i in self.unpaid_invoices)

    def get_contact(self, store_name: str) -> Contact | None:
        """Look up a contact by store name with fuzzy matching."""
        return _lookup_contact(store_name, self.contacts_by_name)

    def print_summary(self) -> None:
        """Print a human-readable summary of what was loaded."""
        unique_locations = {i.store_name for i in self.invoices}
        tier_counts: dict[str, int] = {}
        for inv in self.invoices:
            tier_counts[inv.tier.value] = tier_counts.get(inv.tier.value, 0) + 1

        skip_counts: dict[str, int] = {}
        for inv in self.invoices:
            if inv.skip_reason is not None:
                label = inv.skip_reason.value
                skip_counts[label] = skip_counts.get(label, 0) + 1

        print("=" * 65)
        print("  AR Email Automation -- Data Load Summary")
        print("=" * 65)
        print(f"  Source file       : {self.source_file or '(bytes buffer)'}")
        print(f"  Overdue sheet     : {self.overdue_sheet_used}")
        print(f"  Rows scanned      : {self.total_rows_scanned}")
        print(f"  Empty rows skipped: {self.empty_rows_skipped}")
        print("-" * 65)
        print(f"  Total invoices    : {len(self.invoices)}")
        print(f"  Unique locations  : {len(unique_locations)}")
        print(f"  Paid invoices     : {len(self.paid_invoices)}")
        print(f"  Unpaid invoices   : {len(self.unpaid_invoices)}")
        print(f"  Actionable (sendable): {len(self.actionable_invoices)}")
        print(f"  Total AR          : ${self.total_ar:,.2f}")
        print(f"  Unpaid AR         : ${self.unpaid_ar:,.2f}")
        print("-" * 65)
        print(f"  Contacts loaded   : {len(self.contacts)}")
        matched = len(self.matched_locations)
        total_locs = matched + len(self.unmatched_locations)
        print(f"  Contact match rate: {self.match_rate:.1%}  "
              f"({matched}/{total_locs})")
        if self.unmatched_locations:
            print(f"  Unmatched stores  : {', '.join(self.unmatched_locations)}")
        print("-" * 65)
        print("  Tier distribution:")
        for tier in Tier:
            count = tier_counts.get(tier.value, 0)
            if count:
                print(f"    {tier.value:<22s}: {count}")
        if skip_counts:
            print("-" * 65)
            print("  Skip reasons:")
            for reason, count in sorted(skip_counts.items(),
                                        key=lambda x: -x[1]):
                print(f"    {reason:<45s}: {count}")
        if self.warnings:
            print("-" * 65)
            print(f"  Warnings ({len(self.warnings)}):")
            for w in self.warnings[:20]:
                print(f"    - {w}")
            if len(self.warnings) > 20:
                print(f"    ... and {len(self.warnings) - 20} more")
        print("=" * 65)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_workbook(
    source: Union[str, Path, IO[bytes]],
    *,
    overdue_sheet: str | None = None,
) -> LoadResult:
    """Load invoices and contacts from an AR overdue XLSX workbook.

    Parameters
    ----------
    source:
        File path (``str`` or ``Path``) or a readable bytes buffer
        (``io.BytesIO``).
    overdue_sheet:
        Explicit sheet name to use for invoice data.  When *None* the
        loader auto-detects the most recent ``Overdue {M}-{D}`` sheet.

    Returns
    -------
    LoadResult
        Container with parsed invoices, contacts, matching diagnostics,
        and a ``print_summary()`` helper.
    """
    result = LoadResult()

    # ------------------------------------------------------------------
    # 1. Open workbook
    # ------------------------------------------------------------------
    wb = _open_workbook(source)
    if isinstance(source, (str, Path)):
        result.source_file = str(source)

    # ------------------------------------------------------------------
    # 2. Select the overdue sheet
    # ------------------------------------------------------------------
    if overdue_sheet:
        if overdue_sheet not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{overdue_sheet}' not found.  "
                f"Available: {wb.sheetnames}"
            )
        ws_overdue = wb[overdue_sheet]
        result.overdue_sheet_used = overdue_sheet
    else:
        ws_overdue, sheet_name = _find_most_recent_overdue_sheet(wb)
        result.overdue_sheet_used = sheet_name

    logger.info("Using overdue sheet: %s", result.overdue_sheet_used)

    # ------------------------------------------------------------------
    # 3. Parse invoices
    # ------------------------------------------------------------------
    invoices, scan_meta = _parse_invoices(ws_overdue)
    result.total_rows_scanned = scan_meta["rows_scanned"]
    result.empty_rows_skipped = scan_meta["empty_rows"]
    result.warnings.extend(scan_meta["warnings"])
    result.invoices = invoices

    # ------------------------------------------------------------------
    # 4. Parse contacts from Managers sheet
    # ------------------------------------------------------------------
    if _MANAGERS_SHEET in wb.sheetnames:
        contacts, contact_warnings = _parse_contacts(wb[_MANAGERS_SHEET])
        result.contacts = contacts
        result.warnings.extend(contact_warnings)

        # Build lookup dict keyed on normalized retailer name
        for c in contacts:
            norm = _normalize_name(c.store_name)
            result.contacts_by_name[norm] = c
    else:
        result.warnings.append(
            f"Sheet '{_MANAGERS_SHEET}' not found -- contact data unavailable"
        )

    # ------------------------------------------------------------------
    # 5. Match invoice locations to contacts
    # ------------------------------------------------------------------
    seen_locations: set[str] = set()
    for inv in result.invoices:
        if inv.store_name in seen_locations:
            continue
        seen_locations.add(inv.store_name)

        contact = _lookup_contact(inv.store_name, result.contacts_by_name)
        if contact is not None:
            result.matched_locations.append(inv.store_name)
        else:
            result.unmatched_locations.append(inv.store_name)
            logger.warning("No contact match for location: %s", inv.store_name)

    wb.close()
    return result


def load_contacts_only(
    source: Union[str, Path, IO[bytes]],
) -> list[Contact]:
    """Load only the Managers (contact) sheet -- useful for testing."""
    wb = _open_workbook(source)
    if _MANAGERS_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{_MANAGERS_SHEET}' not found in workbook.")
    contacts, _ = _parse_contacts(wb[_MANAGERS_SHEET])
    wb.close()
    return contacts


def lookup_contact(
    location: str,
    contacts_by_name: dict[str, Contact],
) -> Contact | None:
    """Public wrapper around the fuzzy-match contact lookup."""
    return _lookup_contact(location, contacts_by_name)


# ---------------------------------------------------------------------------
# Workbook opening
# ---------------------------------------------------------------------------

def _open_workbook(source: Union[str, Path, IO[bytes]]) -> Workbook:
    """Open an openpyxl Workbook from a file path or bytes buffer."""
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {path}")
        logger.info("Opening XLSX file: %s", path)
        return openpyxl.load_workbook(path, data_only=True, read_only=False)

    # Bytes buffer (BytesIO or similar)
    logger.info("Opening XLSX from bytes buffer")
    return openpyxl.load_workbook(source, data_only=True, read_only=False)


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def _find_most_recent_overdue_sheet(wb: Workbook) -> tuple[Worksheet, str]:
    """Return the most recent ``Overdue M-D`` sheet, or fall back to ``Data``.

    Sheets are sorted by (month, day) descending to pick the freshest snapshot.
    """
    candidates: list[tuple[int, int, str]] = []

    for name in wb.sheetnames:
        m = _OVERDUE_SHEET_PATTERN.match(name)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            candidates.append((month, day, name))

    if candidates:
        candidates.sort(reverse=True)
        best = candidates[0]
        logger.info(
            "Auto-detected most recent overdue sheet: %s (from %d candidates)",
            best[2], len(candidates),
        )
        return wb[best[2]], best[2]

    # Fallback to "Data" sheet
    if "Data" in wb.sheetnames:
        logger.warning(
            "No 'Overdue' snapshot sheets found; falling back to 'Data' sheet"
        )
        return wb["Data"], "Data"

    raise ValueError(
        "No overdue sheet found.  Expected 'Overdue M-D' or 'Data'.  "
        f"Available sheets: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Invoice parsing
# ---------------------------------------------------------------------------

def _parse_invoices(ws: Worksheet) -> tuple[list[Invoice], dict]:
    """Parse invoice rows from the overdue / data worksheet.

    Returns ``(invoices, metadata_dict)`` where metadata contains
    ``rows_scanned``, ``empty_rows``, and ``warnings``.
    """
    warnings: list[str] = []
    invoices: list[Invoice] = []

    # Build header map from row 1
    header_map = _build_header_map(ws, _OVERDUE_HEADERS)

    # Validate required columns exist
    missing_required = []
    for key in ("order_no", "location", "total_due"):
        if key not in header_map:
            missing_required.append(key)
    if missing_required:
        raise ValueError(
            f"Required columns not found in overdue sheet: {missing_required}.  "
            f"Header row: {[cell.value for cell in ws[1]]}"
        )

    rows_scanned = 0
    empty_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        rows_scanned += 1

        # -- Order No: if null this is a padding row --
        order_no_raw = _cell_value(row, header_map, "order_no")
        if order_no_raw is None or str(order_no_raw).strip() in _NULL_SIGNALS:
            empty_rows += 1
            continue

        # Parse Order No as integer then store as string (models.py uses str)
        try:
            order_no_int = int(float(order_no_raw))
            invoice_number = str(order_no_int)
        except (ValueError, TypeError):
            warnings.append(
                f"Row {row[0].row}: cannot parse Order No "
                f"'{order_no_raw}' -- skipping"
            )
            empty_rows += 1
            continue

        # -- Location --
        location_raw = _cell_value(row, header_map, "location")
        store_name = _clean_str(location_raw)
        if not store_name:
            warnings.append(
                f"Row {row[0].row}: Order {invoice_number} has empty "
                f"Location -- skipping"
            )
            empty_rows += 1
            continue

        # -- Financial --
        amount = _parse_currency(
            _cell_value(row, header_map, "total_due"), default=0.0
        )

        # -- Dates & aging --
        due_date = _parse_date(
            _cell_value(row, header_map, "due_date"),
            f"Row {row[0].row} Due Date", warnings,
        )
        days_past_due = _parse_int(
            _cell_value(row, header_map, "days_over"), default=0
        )

        # -- Status enum --
        status_raw = _clean_str_or_none(
            _cell_value(row, header_map, "status")
        )
        status = _parse_invoice_status(status_raw)

        # -- Boolean flags --
        paid = _parse_bool(_cell_value(row, header_map, "paid"))
        email_sent = _parse_bool(_cell_value(row, header_map, "email_sent"))
        called = _parse_bool(_cell_value(row, header_map, "called"))
        made_contact = _parse_bool(_cell_value(row, header_map, "made_contact"))

        # -- Personnel --
        account_manager = _clean_str(
            _cell_value(row, header_map, "account_mgr")
        )
        account_manager_phone = _clean_str(
            _cell_value(row, header_map, "am_phone")
        )
        sales_rep = _clean_str(_cell_value(row, header_map, "rep"))

        # -- Notes --
        notes = _clean_str(
            _cell_value(row, header_map, "notes")
        )
        follow_up_date = _parse_date(
            _cell_value(row, header_map, "follow_up_date"),
            f"Row {row[0].row} F/U Date", warnings,
        )

        # Build the Invoice object.
        # __post_init__ will auto-assign tier and skip_reason.
        inv = Invoice(
            invoice_number=invoice_number,
            store_name=store_name,
            amount=amount,
            due_date=due_date,
            days_past_due=days_past_due,
            status=status,
            paid=paid,
            email_sent=email_sent,
            called=called,
            made_contact=made_contact,
            account_manager=account_manager,
            account_manager_phone=account_manager_phone,
            sales_rep=sales_rep,
            notes=notes,
            follow_up_date=follow_up_date,
        )

        if inv.amount == 0.0 and _cell_value(row, header_map, "total_due") is not None:
            raw_val = _cell_value(row, header_map, "total_due")
            warnings.append(
                f"Row {row[0].row}: Order {invoice_number} Total Due "
                f"parsed as $0.00 from '{raw_val}'"
            )

        invoices.append(inv)

    meta = {
        "rows_scanned": rows_scanned,
        "empty_rows": empty_rows,
        "warnings": warnings,
    }
    logger.info(
        "Parsed %d invoices from %d rows (%d empty/padding)",
        len(invoices), rows_scanned, empty_rows,
    )
    return invoices, meta


# ---------------------------------------------------------------------------
# Contact parsing
# ---------------------------------------------------------------------------

def _parse_contacts(ws: Worksheet) -> tuple[list[Contact], list[str]]:
    """Parse the Managers sheet into ``Contact`` objects.

    Multi-line POC fields are split on newline and individual entries
    are cleaned.  The *primary* contact is chosen by preferring entries
    with ``(AP)`` or ``Accounting`` in the title, falling back to the
    first listed contact.
    """
    warnings: list[str] = []
    contacts: list[Contact] = []

    header_map = _build_header_map(ws, _MANAGERS_HEADERS)

    if "retailer_name" not in header_map:
        warnings.append(
            "Managers sheet: 'Retailer Name (DBA)' column not found.  "
            f"Header row: {[cell.value for cell in ws[1]]}"
        )
        return contacts, warnings

    for row in ws.iter_rows(min_row=2, values_only=False):
        name_raw = _cell_value(row, header_map, "retailer_name")
        retailer_name = _clean_str(name_raw)
        if not retailer_name:
            continue  # skip empty / padding rows

        # -- Account Manager --
        account_manager = _clean_str(
            _cell_value(row, header_map, "account_mgr")
        )
        account_manager_phone = _clean_str(
            _cell_value(row, header_map, "am_phone")
        )

        # -- POC Name & Title (multi-line) --
        poc_name_raw = _clean_str_or_none(
            _cell_value(row, header_map, "poc_name")
        )
        all_contacts_parsed: list[dict[str, str]] = []
        primary_name = ""
        primary_role = ""

        if poc_name_raw:
            all_contacts_parsed = _parse_poc_names(poc_name_raw)
            primary = _select_primary_contact(all_contacts_parsed)
            primary_name = primary.get("full_name", "")
            primary_role = primary.get("title", "")

        # -- POC Email (multi-line) --
        poc_email_raw = _clean_str_or_none(
            _cell_value(row, header_map, "poc_email")
        )
        all_emails = _parse_emails(poc_email_raw) if poc_email_raw else []

        # Select primary email: prefer ap/accounting/invoices addresses
        primary_email = _select_primary_email(all_emails)

        # -- POC Phone (multi-line) --
        poc_phone_raw = _clean_str_or_none(
            _cell_value(row, header_map, "poc_phone")
        )
        all_phones = _parse_phones(poc_phone_raw) if poc_phone_raw else []
        primary_phone = all_phones[0] if all_phones else ""

        contact = Contact(
            store_name=retailer_name,
            email=primary_email,
            phone=primary_phone,
            contact_name=primary_name,
            role=primary_role,
            all_emails=all_emails,
            all_phones=all_phones,
            all_contacts=all_contacts_parsed,
            account_manager=account_manager,
            account_manager_phone=account_manager_phone,
        )

        contacts.append(contact)

    logger.info("Parsed %d contact records from Managers sheet", len(contacts))
    return contacts, warnings


# ---------------------------------------------------------------------------
# POC field parsers
# ---------------------------------------------------------------------------

def _parse_poc_names(raw: str) -> list[dict[str, str]]:
    """Parse a multi-line POC Name & Title field.

    Each line may be formatted as:
      - ``"FirstName LastName"``
      - ``"FirstName LastName (Title)"``   e.g. ``"Emily Stratakos (AP)"``
      - ``"FirstName LastName - Title"``   e.g. ``"Jo-Anne Rainone - Accounting"``

    Returns a list of dicts: ``{"full_name", "first_name", "title"}``.
    """
    results: list[dict[str, str]] = []
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue

        title = ""
        name = line

        # Parenthesized title: "Emily Stratakos (AP)"
        paren_match = re.match(r"^(.+?)\s*\(([^)]+)\)\s*$", line)
        if paren_match:
            name = paren_match.group(1).strip()
            title = paren_match.group(2).strip()
        else:
            # Dash-separated title: "Jo-Anne Rainone - Accounting"
            # Only split on " - " (with spaces) to avoid splitting
            # hyphenated names like "Jo-Anne".
            dash_parts = line.split(" - ", 1)
            if len(dash_parts) == 2:
                candidate_name = dash_parts[0].strip()
                candidate_title = dash_parts[1].strip()
                if candidate_title and not candidate_title[0].isdigit():
                    name = candidate_name
                    title = candidate_title

        first_name = name.split()[0] if name.split() else name
        results.append({
            "full_name": name,
            "first_name": first_name,
            "title": title,
            "name": name,  # alias used in Contact.all_contacts
        })

    return results


def _select_primary_contact(parsed_names: list[dict[str, str]]) -> dict[str, str]:
    """Select the best contact for AR emails.

    Priority:
    1. Contact with ``(AP)`` in title
    2. Contact with ``Accounting`` / ``Finance`` / ``Billing`` in title
    3. First listed contact
    """
    for entry in parsed_names:
        title = (entry.get("title") or "").lower()
        for kw in _AR_TITLE_KEYWORDS:
            if kw in title:
                return entry

    return parsed_names[0] if parsed_names else {
        "full_name": "", "first_name": "", "title": "",
    }


def _parse_emails(raw: str) -> list[str]:
    """Parse multi-line email field into de-duplicated, lowercased list.

    Handles newline-delimited and sometimes comma- or semicolon-delimited
    entries.  Performs basic ``@`` validation.
    """
    emails: list[str] = []
    seen: set[str] = set()

    for part in re.split(r"[\n,;]+", raw):
        email = part.strip().lower()
        if not email:
            continue
        # Basic email validation
        if "@" in email and "." in email.split("@")[-1]:
            if email not in seen:
                emails.append(email)
                seen.add(email)
        else:
            logger.debug("Skipping invalid email token: %r", part.strip())

    return emails


def _select_primary_email(emails: list[str]) -> str:
    """Pick the best email for AR correspondence.

    Priority:
    1. Address containing ``ap@``, ``accounting@``, ``invoices@``
    2. First address in the list
    """
    ar_prefixes = ["ap@", "accounting@", "invoices@", "billing@"]
    for email in emails:
        for prefix in ar_prefixes:
            if prefix in email:
                return email
    return emails[0] if emails else ""


def _parse_phones(raw: str) -> list[str]:
    """Parse multi-line phone field.

    Lines may be formatted as ``"Name - (xxx) xxx-xxxx"`` or just a
    phone number.  Returns the full line (preserving the label).
    """
    phones: list[str] = []
    for line in raw.split("\n"):
        line = line.strip()
        if line:
            phones.append(line)
    return phones


# ---------------------------------------------------------------------------
# Contact lookup / fuzzy matching
# ---------------------------------------------------------------------------

def _normalize_name(name: str) -> str:
    """Normalize a retailer name for fuzzy matching.

    Lowercases, strips trailing punctuation and extra whitespace.
    """
    return re.sub(r"[.\s]+$", "", name.strip().lower())


def _lookup_contact(
    location: str,
    contacts_by_name: dict[str, Contact],
) -> Contact | None:
    """Look up a contact by store location with fuzzy fallback.

    Matching strategy:
    1. Exact normalized match.
    2. Substring containment -- handles trailing-period differences
       like "HUB Dispensary" vs "HUB Dispensary." and case differences
       like "Transcend Wellness" vs "Transcend wellness".
    """
    norm = _normalize_name(location)

    # 1. Exact normalized match
    if norm in contacts_by_name:
        return contacts_by_name[norm]

    # 2. Substring containment
    for key, contact in contacts_by_name.items():
        if norm in key or key in norm:
            logger.info(
                "Fuzzy match: '%s' -> '%s'", location, contact.store_name
            )
            return contact

    return None


# ---------------------------------------------------------------------------
# Header map builder
# ---------------------------------------------------------------------------

def _build_header_map(
    ws: Worksheet,
    header_spec: dict[str, list[str]],
) -> dict[str, int]:
    """Map logical field names to 0-based column indices.

    Reads row 1 of the worksheet and matches each header cell against
    the known aliases in *header_spec*.
    """
    header_map: dict[str, int] = {}

    row1_values: list[str | None] = []
    for cell in ws[1]:
        val = cell.value
        row1_values.append(str(val).strip() if val is not None else None)

    for logical_name, aliases in header_spec.items():
        for idx, header_text in enumerate(row1_values):
            if header_text is None:
                continue
            for alias in aliases:
                if header_text.lower() == alias.lower():
                    header_map[logical_name] = idx
                    break
            if logical_name in header_map:
                break

    found = list(header_map.keys())
    logger.debug("Header map (%d/%d): %s",
                 len(found), len(header_spec), found)
    return header_map


# ---------------------------------------------------------------------------
# Cell reading helpers
# ---------------------------------------------------------------------------

def _cell_value(row, header_map: dict[str, int], field_name: str):
    """Safely read a cell value by logical field name.

    Returns ``None`` if the field is not in the header map or the cell
    index is beyond the row length.
    """
    idx = header_map.get(field_name)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx].value


# ---------------------------------------------------------------------------
# Data cleaning / type coercion helpers
# ---------------------------------------------------------------------------

def _clean_str(val) -> str:
    """Convert a cell value to a stripped string.  None becomes ``""``."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in _NULL_SIGNALS else s


def _clean_str_or_none(val) -> str | None:
    """Convert a cell value to a stripped string, returning None for nullish."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s in _NULL_SIGNALS else s


def _parse_bool(val) -> bool:
    """Parse a boolean cell value.

    Handles ``True``, ``False``, ``"TRUE"``, ``"FALSE"``, ``1``, ``0``.
    """
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes")


def _parse_int(val, default: int = 0) -> int:
    """Parse an integer cell value, returning *default* on failure."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _parse_currency(val, default: float = 0.0) -> float:
    """Parse a dollar-amount cell value.

    Handles:
    - Numeric floats from openpyxl (most common path).
    - Strings like ``"$1,234.56"`` or ``"1234.56"``.
    - Parenthesized negatives: ``"($500.00)"``.
    """
    if val is None:
        return default

    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if not s or s in _NULL_SIGNALS:
        return default

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").strip()

    try:
        amount = float(s)
        return -amount if negative else amount
    except ValueError:
        return default


def _parse_date(val, context: str, warnings: list[str]) -> date | None:
    """Parse a date cell value.

    openpyxl typically returns ``datetime`` objects for date-typed cells.
    Also handles Excel serial date numbers and common string formats.
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # Numeric -- might be an Excel serial date
    if isinstance(val, (int, float)):
        try:
            serial = int(val)
            if 40000 < serial < 50000:
                base = datetime(1899, 12, 30)
                return (base + timedelta(days=serial)).date()
        except (ValueError, OverflowError):
            pass

    # String date formats
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y",
                "%Y-%m-%dT%H:%M:%S", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    warnings.append(f"{context}: could not parse date '{val}'")
    return None


def _parse_invoice_status(raw: str | None) -> InvoiceStatus:
    """Map a raw status string to the ``InvoiceStatus`` enum.

    Returns ``InvoiceStatus.NONE`` for null / unrecognised values.
    """
    if raw is None:
        return InvoiceStatus.NONE

    # Try exact match first
    for member in InvoiceStatus:
        if member.value == raw:
            return member

    # Case-insensitive fallback
    raw_lower = raw.lower().strip()
    for member in InvoiceStatus:
        if member.value.lower() == raw_lower:
            return member

    logger.warning("Unknown invoice status value: '%s' -- defaulting to NONE", raw)
    return InvoiceStatus.NONE


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    from collections import Counter

    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s | %(name)s | %(message)s",
    )

    # Default path relative to this file
    default_path = (
        Path(__file__).resolve().parent.parent
        / "data"
        / "NY Account Receivables_Overdue.xlsx"
    )
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else str(default_path)

    print(f"\nLoading: {xlsx_path}\n")
    result = load_workbook(xlsx_path)
    result.print_summary()

    # -- Extra diagnostics --

    print("\n--- Actionable Invoice Sample (first 5) ---")
    for inv in result.actionable_invoices[:5]:
        contact = result.get_contact(inv.store_name)
        poc_email = contact.email if contact else "(no email)"
        poc_name = contact.first_name if contact else "(no contact)"
        print(
            f"  #{inv.invoice_number}  {inv.store_name:<35s}  "
            f"${inv.amount:>9,.2f}  {inv.days_past_due:>4d}d  "
            f"{inv.tier.value:<22s}  "
            f"{poc_name} <{poc_email}>"
        )

    # Multi-invoice locations
    loc_counts = Counter(inv.store_name for inv in result.invoices)
    multi = {loc: cnt for loc, cnt in loc_counts.items() if cnt > 1}
    if multi:
        print(f"\n--- Multi-Invoice Locations ({len(multi)}) ---")
        for loc, cnt in sorted(multi.items(), key=lambda x: -x[1]):
            total = sum(i.amount for i in result.invoices
                        if i.store_name == loc)
            print(f"  {loc:<35s}  {cnt} invoices  ${total:>9,.2f}")
